import csv
import io
import re
from datetime import date, datetime
from uuid import UUID

from fastapi import UploadFile
from pydantic import ValidationError

from app.core.exceptions import AppException
from app.models.user import User
from app.schemas.blog import BlogPostCreate
from app.schemas.bulk import BulkImportReport
from app.schemas.category import CategoryCreate
from app.schemas.coupon import CouponCreate
from app.schemas.product import ProductCreate
from app.services.blog_service import BlogService
from app.services.category_service import CategoryService
from app.services.coupon_service import CouponService
from app.services.product_service import ProductService

ALLOWED_EXTENSIONS = {".csv", ".xlsx"}

TRUTHY = {"1", "true", "yes", "y", "on"}
FALSY = {"0", "false", "no", "n", "off", ""}

IMAGE_COLUMN_RE = re.compile(r"^image(\d+)$")


def _normalize_header(header: str) -> str:
    return header.strip().lower().replace(" ", "_")


def _require(value: str | None, field: str) -> str:
    if value is None or not value.strip():
        raise ValueError(f"{field} is required")
    return value.strip()


def _optional(value: str | None) -> str | None:
    if value is None or not value.strip():
        return None
    return value.strip()


def _split(value: str | None) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in re.split(r"[,|]", value) if part.strip()]


def _as_bool(value: str | None) -> bool:
    normalized = (value or "").strip().lower()
    if normalized in FALSY:
        return False
    return normalized in TRUTHY


def _as_float(value: str | None, field: str) -> float:
    if value is None or value == "":
        raise ValueError(f"{field} is required")
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        raise ValueError(f"{field} must be a number, got \"{value}\"")


def _optional_float(value: str | None) -> float | None:
    if value is None or value == "":
        return None
    return _as_float(value, "compare_at_price")


def _as_int(value: str | None, field: str) -> int:
    if value is None or value == "":
        raise ValueError(f"{field} is required")
    try:
        return int(float(str(value).replace(",", "").strip()))
    except ValueError:
        raise ValueError(f"{field} must be a number, got \"{value}\"")


def _as_datetime(value: str | None) -> datetime | None:
    if not value or not value.strip():
        return None
    value = value.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    raise ValueError(f"Could not parse date \"{value}\" (use YYYY-MM-DD)")


def _parse_images(row: dict[str, str]) -> list[str]:
    images: list[str] = []
    seen: set[str] = set()

    def add(url: str) -> None:
        url = (url or "").strip()
        if url and url not in seen:
            seen.add(url)
            images.append(url)

    for key, value in row.items():
        if IMAGE_COLUMN_RE.match(key):
            add(value)

    for url in _split(row.get("images")):
        add(url)

    return images


def _parse_specs(value: str | None) -> list[dict]:
    specs: list[dict] = []
    for part in _split(value):
        if ":" not in part:
            raise ValueError(f"Spec \"{part}\" must be in 'Label: value' format")
        label, _, spec_value = part.partition(":")
        label = label.strip()
        if label:
            specs.append({"label": label, "value": spec_value.strip()})
    return specs


def _message(exc: Exception) -> str:
    if isinstance(exc, AppException):
        return str(exc.message)
    if isinstance(exc, ValidationError):
        first = exc.errors()[0] if exc.errors() else {}
        loc = ".".join(str(part) for part in first.get("loc", []))
        msg = first.get("msg", "")
        return f"{loc}: {msg}" if loc else msg
    return str(exc)


def _report(total: int, errors: list[tuple[int, str]]) -> BulkImportReport:
    return BulkImportReport(
        total=total,
        imported=total - len(errors),
        failed=len(errors),
        errors=[{"row": row, "error": message} for row, message in errors],
    )


class BulkImportService:
    def __init__(
        self,
        product_service: ProductService,
        category_service: CategoryService,
        coupon_service: CouponService,
        blog_service: BlogService,
    ):
        self.product_service = product_service
        self.category_service = category_service
        self.coupon_service = coupon_service
        self.blog_service = blog_service

    # -------------------------
    # FILE PARSING
    # -------------------------
    def parse_upload(self, file: UploadFile) -> list[dict[str, str]]:
        filename = (file.filename or "").lower()
        extension = "." + filename.rsplit(".", 1)[-1] if "." in filename else ""
        if extension not in ALLOWED_EXTENSIONS:
            raise AppException(
                status_code=400,
                error_code="UNSUPPORTED_FILE",
                message="Unsupported file type. Upload a .csv or .xlsx file.",
            )

        content = file.file.read()

        if extension == ".xlsx":
            return self._parse_xlsx(content)
        return self._parse_csv(content)

    def _parse_csv(self, content: bytes) -> list[dict[str, str]]:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows: list[dict[str, str]] = []
        for raw in reader:
            rows.append(
                {_normalize_header(key): (value or "").strip() for key, value in raw.items()}
            )
        return rows

    def _parse_xlsx(self, content: bytes) -> list[dict[str, str]]:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)

        try:
            headers = next(rows)
        except StopIteration:
            return []

        headers = [_normalize_header(str(header)) if header is not None else "" for header in headers]

        parsed: list[dict[str, str]] = []
        for values in rows:
            row: dict[str, str] = {}
            for header, value in zip(headers, values):
                if header:
                    if value is None:
                        row[header] = ""
                    elif isinstance(value, (datetime, date)):
                        row[header] = str(value)
                    else:
                        row[header] = str(value).strip()
            if any(row.values()):
                parsed.append(row)
        return parsed

    # -------------------------
    # PRODUCTS
    # -------------------------
    def import_products(self, file: UploadFile, actor: User) -> BulkImportReport:
        rows = self.parse_upload(file)
        errors: list[tuple[int, str]] = []

        for index, row in enumerate(rows, start=2):
            try:
                category_id = self._resolve_category(row.get("category"))
                images = _parse_images(row)

                product = ProductCreate(
                    name=_require(row.get("name"), "name"),
                    short_description=_require(row.get("short_description"), "short_description"),
                    description=_require(row.get("description"), "description"),
                    category_id=category_id,
                    price=_as_float(row.get("price"), "price"),
                    compare_at_price=_optional_float(row.get("compare_at_price")),
                    currency=row.get("currency") or "NGN",
                    unit=row.get("unit") or "head",
                    stock=_as_int(row.get("stock") or "0", "stock"),
                    origin=_optional(row.get("origin")),
                    farm=_optional(row.get("farm")),
                    images=images,
                    specs=_parse_specs(row.get("specs")),
                    tags=_split(row.get("tags")),
                    badges=[badge for badge in _split(row.get("badges")) if badge],
                    status=(row.get("status") or "published").lower(),
                )

                self.product_service.create_product(product, actor)
            except (AppException, ValueError, ValidationError) as exc:
                errors.append((index, _message(exc)))

        return _report(len(rows), errors)

    def _resolve_category(self, value: str | None) -> UUID:
        name = (value or "").strip()
        if not name:
            raise ValueError("category is required")

        category = (
            self.category_service.get_category_by_slug(name)
            or self.category_service.get_category_by_name(name)
        )
        if not category:
            raise ValueError(f"category \"{name}\" not found")
        return category.id

    # -------------------------
    # CATEGORIES
    # -------------------------
    def import_categories(self, file: UploadFile, actor: User) -> BulkImportReport:
        rows = self.parse_upload(file)
        errors: list[tuple[int, str]] = []

        for index, row in enumerate(rows, start=2):
            try:
                category = CategoryCreate(
                    name=_require(row.get("name"), "name"),
                    slug=_optional(row.get("slug")),
                    short_description=_require(row.get("short_description"), "short_description"),
                    description=_require(row.get("description"), "description"),
                    image=_optional(row.get("image")),
                    emoji=_optional(row.get("emoji")),
                    accent=_optional(row.get("accent")),
                    featured=_as_bool(row.get("featured")),
                )

                self.category_service.create_category(category, actor)
            except (AppException, ValueError, ValidationError) as exc:
                errors.append((index, _message(exc)))

        return _report(len(rows), errors)

    # -------------------------
    # COUPONS
    # -------------------------
    def import_coupons(self, file: UploadFile, actor: User) -> BulkImportReport:
        rows = self.parse_upload(file)
        errors: list[tuple[int, str]] = []

        for index, row in enumerate(rows, start=2):
            try:
                coupon = CouponCreate(
                    code=_require(row.get("code"), "code"),
                    type=_require(row.get("type"), "type"),
                    value=_as_float(row.get("value"), "value"),
                    min_order=_optional_float(row.get("min_order")) or 0,
                    description=_optional(row.get("description")),
                    expires_at=_as_datetime(row.get("expires_at")),
                )

                self.coupon_service.create_coupon(coupon, actor)
            except (AppException, ValueError, ValidationError) as exc:
                errors.append((index, _message(exc)))

        return _report(len(rows), errors)

    # -------------------------
    # BLOG POSTS
    # -------------------------
    def import_blog_posts(self, file: UploadFile, actor: User) -> BulkImportReport:
        rows = self.parse_upload(file)
        errors: list[tuple[int, str]] = []

        for index, row in enumerate(rows, start=2):
            try:
                images = _parse_images(row)
                content_raw = _require(row.get("content"), "content")
                paragraphs = [part.strip() for part in re.split(r"\n\s*\n|\|", content_raw) if part.strip()]

                post = BlogPostCreate(
                    title=_require(row.get("title"), "title"),
                    excerpt=_require(row.get("excerpt"), "excerpt"),
                    content=paragraphs,
                    category=_require(row.get("category"), "category"),
                    tags=_split(row.get("tags")),
                    featured=_as_bool(row.get("featured")),
                    cover_image=_optional(row.get("cover_image")),
                    images=images,
                )

                self.blog_service.create_post(post, actor)
            except (AppException, ValueError, ValidationError) as exc:
                errors.append((index, _message(exc)))

        return _report(len(rows), errors)
